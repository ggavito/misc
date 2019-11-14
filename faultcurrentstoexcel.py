# TXT to Excel files
# FAULT study
# Written by: Garet Gavito
# made for use by electrical engineers
#
# To install needed library, have up to date python 3 installed
# type cmd into start menu to open windows command prompt.
# in command prompt, type python -m pip install openpyxl

# to use this, simply run it in the same folder as SKM outputs
# outputs excel file, columns are in the order of provided file names

# retrieves list of fault currents based on lines of text that don't change
# at least in this version of SKM. there is probably a cleaner way to do this
# but this is quick, easy, and it works.

from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl
import re

voltageRange = []
filenames = input('\nEnter txt file(s) to read separated by commas (E.G. NORMAL_ON,NORMAL_OFF,FUTURE_ON): ')
filenames = filenames.split(',') # split input into list of strings
highside = ''
closest = input('Name closest turbine (e.g. C01): ')
furthest = input('Name furthest turbine (e.g. E10): ')

def readRpt(filename):
    # find and open the report file output from SKM and interpret it
    # high side voltage needs to be set manually right now
    filenamerpt = filename + ".rpt"
    f = open(filenamerpt, 'r')
    search = f.readlines()
    prevLine = ""

    # look for high-side bus voltage in the first 400 lines of the document
    # 400 lines is a guess and may need to be adjusted for particularly tiny or large wind farms
    for i in range(1,400):
        highsideMatch = (re.findall('\d{3}kV\-Bus',search[i]))
        if highsideMatch:
            try:
                highside = highsideMatch[0]
            except:
                print("Error finding highside voltage")
    print(highside)

    faultsDict = {highside:[],'34.5kV-Bus':[],'HS-' + closest:[],
                  'HS-' + furthest:[],'LS-' + closest:[],
                  'LS-' + furthest:[]} # fault current dictionary
    buslist = [highside,'34.5kV-Bus','HS-' + closest,
                  'HS-' + furthest,'LS-' + closest,
                  'LS-' + furthest] # list of bus names
    
    # search function
    # retrieves fault currents into dict
    # order is 3PH, SLG, LL, LLG for each bus it finds
    for i, line in enumerate(search):
        for bus in buslist:
            if bus in prevLine and "INI. SYM. RMS FAULT CURRENT:  " in line:
                fault = re.findall("\d+\.\d*", line)
                faultsDict[bus].append(fault[0])
        prevLine = line                     
    f.close()
    return faultsDict, buslist

def toExcel(dictList, filename, col):
    # put values into excel spreadsheet for copy pasting into report
    # overwrites the same spreadsheet each time
    faultsDict = dictList[0]
    buslist = dictList[1]

    destFilename = 'fault_currents.xlsx'
    if col == 1:
        wb = Workbook()
        ws = wb.active
        ws.title = destFilename
    else:
        wb = load_workbook(destFilename)
        ws = wb.active
    startrow = 1
    for bus in buslist:
        for row, value in zip(range(startrow,len(faultsDict)*4+1),faultsDict[bus]):
            _ = ws.cell(column=col,row=row, value='=ROUND(' + value + ',0)')
            startrow = row+1

    wb.save(filename=destFilename)
col = 1
for file in filenames:
    toExcel(readRpt(file),file,col)
    col += 1
print('Search completed successfully')
print('Please find results in fault_currents.xlsx')
